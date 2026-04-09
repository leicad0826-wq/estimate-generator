import streamlit as st
import subprocess, os, shutil, glob, zipfile, re, tempfile
from lxml import etree
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
from PIL import Image
import io

st.set_page_config(page_title="カードラボ「最終見積書」自動作成", page_icon="🌸", layout="centered")

# ============================================================
# スタイル
# ============================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@300;400;500;700&display=swap');

html, body, [class*="css"] { font-family: 'Noto Sans JP', sans-serif; }

.main { background: #fdf6f9; }

.hero {
    background: linear-gradient(135deg, #e8b4c8 0%, #f2d0e0 40%, #c9a4d4 100%);
    border-radius: 20px;
    padding: 36px 32px;
    color: #4a2040;
    margin-bottom: 28px;
    position: relative;
    overflow: hidden;
    border: 1px solid #e8a0c0;
}
.hero::before {
    content: '';
    position: absolute;
    top: -40%;
    right: -15%;
    width: 300px;
    height: 300px;
    background: radial-gradient(circle, rgba(255,255,255,0.3) 0%, transparent 70%);
    border-radius: 50%;
}
.hero h1 { font-size: 1.6rem; font-weight: 700; margin: 0 0 8px 0; letter-spacing: -0.5px; }
.hero p  { font-size: 0.9rem; opacity: 0.7; margin: 0; }

.step-card {
    background: white;
    border-radius: 16px;
    padding: 18px 22px;
    margin-bottom: 14px;
    border: 1px solid #f0d4e4;
    box-shadow: 0 2px 10px rgba(200,120,160,0.08);
}
.step-num {
    display: inline-block;
    background: linear-gradient(135deg, #d4849c, #c9a4d4);
    color: white;
    border-radius: 50%;
    width: 28px;
    height: 28px;
    line-height: 28px;
    text-align: center;
    font-size: 0.85rem;
    font-weight: 700;
    margin-right: 10px;
}
.step-title { font-weight: 600; font-size: 1rem; color: #4a2040; }

.success-box {
    background: linear-gradient(135deg, #fce4ec, #f3e5f5);
    border: 2px solid #e091b5;
    border-radius: 16px;
    padding: 24px 28px;
    margin-top: 20px;
    text-align: center;
}
.success-box b { color: #8e2462; font-size: 1.15rem; }
.success-box small { color: #7b3a6e; }

.warning-box {
    background: #fff8f0;
    border: 1px solid #f5c28a;
    border-radius: 12px;
    padding: 14px 18px;
    font-size: 0.9rem;
    color: #8a5a20;
}

[data-testid="stFileUploader"] {
    border: 2px dashed #e0b0c8 !important;
    border-radius: 14px !important;
    background: #fdf6f9 !important;
}

/* ダウンロードボタン（success-boxと同じサイズ感） */
[data-testid="stDownloadButton"] > button {
    background: linear-gradient(135deg, #d4549c, #a044a0) !important;
    color: white !important;
    border: none !important;
    border-radius: 16px !important;
    padding: 24px 28px !important;
    font-size: 1.5rem !important;
    font-weight: 700 !important;
    box-shadow: 0 4px 16px rgba(180,60,120,0.3) !important;
    transition: all 0.2s !important;
    white-space: pre-line !important;
    line-height: 1.6 !important;
}
[data-testid="stDownloadButton"] > button p,
[data-testid="stDownloadButton"] > button span,
[data-testid="stDownloadButton"] > button div,
[data-testid="stDownloadButton"] > button * {
    font-size: 1.5rem !important;
    font-weight: 700 !important;
}
[data-testid="stDownloadButton"] > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 20px rgba(180,60,120,0.4) !important;
}

/* 生成ボタン */
[data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg, #d4849c, #c9a4d4) !important;
    border: none !important;
    border-radius: 12px !important;
    font-weight: 600 !important;
}
</style>
""", unsafe_allow_html=True)

# ============================================================
# ヘッダー
# ============================================================
st.markdown("""
<div class="hero">
    <h1>🌸 カードラボ「最終見積書」自動作成</h1>
    <p>見積算出表（.xlsb）をアップロードするだけで最終見積書を自動生成します</p>
</div>
""", unsafe_allow_html=True)

# ============================================================
# generate_estimate の関数群（インライン）
# ============================================================
NS_XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
NS_A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'
NS_R   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
NS_PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'
NS_S   = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NS_CT  = 'http://schemas.openxmlformats.org/package/2006/content-types'
SLOTS  = [24, 26, 28, 30, 32, 34, 36, 38, 40]

def parse_betto(text):
    items = []
    if not text: return items
    for line in str(text).strip().split('\n'):
        line = line.strip()
        if not line: continue
        m = re.match(r'^(.+?)\s+(\d+)([^\d\s]+)\s+(\d+)円?$', line)
        if m:
            items.append({'name': m.group(1).strip(), 'qty': int(m.group(2)),
                          'unit': m.group(3), 'price': int(m.group(4))})
        else:
            parts = line.split('\u3000')
            if len(parts) == 3:
                name = parts[0].strip()
                qu = re.match(r'(\d+)(.+)', parts[1].strip())
                price = re.sub(r'[^\d]', '', parts[2])
                if qu and price:
                    items.append({'name': name, 'qty': int(qu.group(1)),
                                  'unit': qu.group(2), 'price': int(price)})
                else:
                    items.append({'name': line, 'qty': 1, 'unit': '式', 'price': 0})
            elif len(parts) == 2:
                name = parts[0].strip()
                price = re.sub(r'[^\d]', '', parts[1])
                if price:
                    items.append({'name': name, 'qty': 1, 'unit': '式', 'price': int(price)})
                else:
                    items.append({'name': line, 'qty': 1, 'unit': '式', 'price': 0})
            else:
                items.append({'name': line, 'qty': 1, 'unit': '式', 'price': 0})
    return items

def read_estimate(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['見積算出表']
    anken = str(ws['D7'].value or '')
    anken_stripped = re.sub(r'^\d{3}', '', anken).strip()
    nouki = ws['D10'].value
    weekdays = ['月', '火', '水', '木', '金', '土', '日']
    if isinstance(nouki, datetime):
        nouki = f"{nouki.strftime('%Y年%m月%d日')}（{weekdays[nouki.weekday()]}）"
    elif isinstance(nouki, (int, float)) and 40000 < nouki < 60000:
        dt = datetime(1899, 12, 30) + timedelta(days=int(nouki))
        nouki = f"{dt.strftime('%Y年%m月%d日')}（{weekdays[dt.weekday()]}）"
    honnohin = ws['D43'].value or 0
    yubi     = ws['K43'].value or 0
    tanka = None
    for r in range(100, 160):
        for c in range(14, 20):
            if ws.cell(row=r, column=c).value == '最終単価':
                tanka = ws.cell(row=r, column=17).value; break
        if tanka: break
    betto_text = None
    for r in range(100, 160):
        for c in range(18, 24):
            if ws.cell(row=r, column=c).value == '別途請求内容':
                betto_text = ws.cell(row=r+1, column=c).value; break
        if betto_text is not None: break
    sku = ws['D15'].value or 1
    insatsu_men = ws['D16'].value or ''
    insatsu_ura = ws['D18'].value or ''
    item_info   = str(ws['D8'].value or '')
    size_match  = re.search(r'(\d+[×x]\d+mm)', item_info)
    if size_match:
        size_label = '枠サイズ'; size_str = size_match.group(1)
    else:
        size_label = 'デザイン本体サイズ'; size_str = item_info
    insatsu_display = f"{insatsu_men}面）{insatsu_ura}" if insatsu_men and insatsu_ura else str(insatsu_ura)
    return dict(anken=anken, anken_stripped=anken_stripped, nouki=str(nouki),
                honnohin=honnohin, yubi=yubi, tanka=tanka,
                betto_items=parse_betto(betto_text), sku=sku,
                size_label=size_label, size_str=size_str, insatsu=insatsu_display)

def get_images_from_sheet(file_path, sheet_name='種別'):
    images = []
    is_xlsb = file_path.lower().endswith('.xlsb')
    with zipfile.ZipFile(file_path) as z:
        if is_xlsb:
            import pyxlsb as _pyxlsb
            with _pyxlsb.open_workbook(file_path) as wb_xlsb:
                sheet_idx = None
                for i, sname in enumerate(wb_xlsb.sheets):
                    if sname == sheet_name:
                        sheet_idx = i; break
            if sheet_idx is None: return images
            sheet_file = f'sheet{sheet_idx + 1}.bin'
            rels_path = f'xl/worksheets/_rels/{sheet_file}.rels'
        else:
            wb_xml  = etree.fromstring(z.read('xl/workbook.xml'))
            wb_rels = etree.fromstring(z.read('xl/_rels/workbook.xml.rels'))
            rid_to_target = {r.get('Id'): r.get('Target') for r in wb_rels}
            sheet_file = None
            for sh in wb_xml.findall(f'{{{NS_S}}}sheets/{{{NS_S}}}sheet'):
                if sh.get('name') == sheet_name:
                    rid = sh.get(f'{{{NS_R}}}id')
                    sheet_file = rid_to_target.get(rid, '').split('/')[-1]
                    break
            if not sheet_file: return images
            rels_path = f'xl/worksheets/_rels/{sheet_file}.rels'
        if rels_path not in z.namelist(): return images
        sheet_rels = etree.fromstring(z.read(rels_path))
        drawing_target = None
        for rel in sheet_rels:
            if 'drawing' in rel.get('Type','') and 'vml' not in rel.get('Target',''):
                drawing_target = rel.get('Target'); break
        if not drawing_target: return images
        drawing_path      = 'xl/' + drawing_target.lstrip('../')
        drawing_rels_path = drawing_path.replace('drawings/','drawings/_rels/') + '.rels'
        if drawing_path not in z.namelist(): return images
        drawing = etree.fromstring(z.read(drawing_path))
        d_rels  = etree.fromstring(z.read(drawing_rels_path)) if drawing_rels_path in z.namelist() else None
        rid_map = {r.get('Id'): r.get('Target') for r in d_rels} if d_rels is not None else {}
        for anchor in drawing:
            fe  = anchor.find(f'{{{NS_XDR}}}from')
            pic = anchor.find(f'{{{NS_XDR}}}pic')
            if fe is None or pic is None: continue
            fc = int(fe.find(f'{{{NS_XDR}}}col').text)
            fr = int(fe.find(f'{{{NS_XDR}}}row').text)
            blip = pic.find(f'.//{{{NS_A}}}blip')
            if blip is None: continue
            rid = blip.get(f'{{{NS_R}}}embed','')
            img_rel  = rid_map.get(rid,'')
            img_path = 'xl/media/' + img_rel.split('/')[-1]
            if img_path in z.namelist():
                images.append((z.read(img_path), fc, fr))
    images.sort(key=lambda x: (x[2], x[1]))
    return [(d,) for d,_,_ in images]

IMG_COL_START = 20   # 画像配置開始列（備考テキスト右側）
IMG_ROW_START = 44   # 画像配置開始行（45行目=0-indexed 44）
EMU_PER_COL   = 181000  # 1列あたり概算EMU（列幅13 chars）
MAX_CX = 18 * EMU_PER_COL           # col20〜col38 = 18列分
MAX_CY = 3300000                     # row44〜row59(45行目〜60行目) 61行目にかからない
IMG_GAP_EMU = 100000                 # オブジェクト間の余白（約1mm）

def extract_objects(img_data):
    """デザイン画像からヘッダーテキストを除去し、本体・台座を個別に切り出す"""
    import numpy as np
    img = Image.open(io.BytesIO(img_data)).convert('RGB')
    arr = np.array(img)
    non_white = ~((arr[:,:,0] > 240) & (arr[:,:,1] > 240) & (arr[:,:,2] > 240))
    row_sums = non_white.sum(axis=1)
    pad = 3

    # ヘッダーテキスト除外: 20行以上の空白ギャップ後のコンテンツ開始行を探す
    gap_count = 0
    body_start = 0
    for r in range(len(row_sums)):
        if row_sums[r] == 0:
            gap_count += 1
        else:
            if gap_count > 20:
                body_start = r
                break
            gap_count = 0

    body_area = non_white[body_start:, :]
    col_sums = body_area.sum(axis=0)

    # 列方向のギャップ検出（30px以上の空白列）でオブジェクト分割
    sparse = col_sums < 5
    in_gap = False
    gaps = []
    gs = 0
    for c in range(len(sparse)):
        if sparse[c] and not in_gap:
            gs = c; in_gap = True
        elif not sparse[c] and in_gap:
            if c - gs > 30:
                gaps.append((gs, c))
            in_gap = False

    content_cols = np.where(col_sums >= 5)[0]
    if len(content_cols) == 0:
        return [(img_data, img.size[0], img.size[1])]

    # オブジェクト領域を収集
    regions = []
    prev = content_cols[0]
    for g_start, g_end in gaps:
        if g_start > prev:
            regions.append((prev, g_start))
        prev = g_end
    regions.append((prev, content_cols[-1] + 1))

    results = []
    for c1, c2 in regions:
        region = body_area[:, c1:c2]
        rs = region.sum(axis=1)
        rr = np.where(rs > 0)[0]
        if len(rr) == 0:
            continue
        r1 = rr[0] + body_start
        r2 = rr[-1] + body_start
        # 小さすぎるオブジェクト（テキスト片等）は除外
        obj_w = c2 - c1
        obj_h = r2 - r1 + 1
        if obj_w < 50 or obj_h < 50:
            continue
        crop = img.crop((max(0, c1-pad), max(0, r1-pad),
                         min(img.width, c2+pad), min(img.height, r2+1+pad)))
        buf = io.BytesIO()
        crop.save(buf, format='PNG')
        results.append((buf.getvalue(), crop.size[0], crop.size[1]))

    return results if results else [(img_data, img.size[0], img.size[1])]

def make_pic_anchor(rid, img_id, img_name, col, row, cx, cy, col_off=0):
    """oneCellAnchor: 開始セル＋EMUサイズ指定（アスペクト比固定）"""
    anchor = etree.Element(f'{{{NS_XDR}}}oneCellAnchor')
    fe = etree.SubElement(anchor, f'{{{NS_XDR}}}from')
    etree.SubElement(fe, f'{{{NS_XDR}}}col').text    = str(col)
    etree.SubElement(fe, f'{{{NS_XDR}}}colOff').text = str(col_off)
    etree.SubElement(fe, f'{{{NS_XDR}}}row').text    = str(row)
    etree.SubElement(fe, f'{{{NS_XDR}}}rowOff').text = '0'
    ext_elem = etree.SubElement(anchor, f'{{{NS_XDR}}}ext')
    ext_elem.set('cx', str(cx)); ext_elem.set('cy', str(cy))
    pic = etree.SubElement(anchor, f'{{{NS_XDR}}}pic')
    nv  = etree.SubElement(pic, f'{{{NS_XDR}}}nvPicPr')
    cNvPr = etree.SubElement(nv, f'{{{NS_XDR}}}cNvPr')
    cNvPr.set('id', str(img_id)); cNvPr.set('name', img_name)
    cNvPicPr = etree.SubElement(nv, f'{{{NS_XDR}}}cNvPicPr')
    locks = etree.SubElement(cNvPicPr, f'{{{NS_A}}}picLocks'); locks.set('noChangeAspect','1')
    bf = etree.SubElement(pic, f'{{{NS_XDR}}}blipFill')
    blip = etree.SubElement(bf, f'{{{NS_A}}}blip'); blip.set(f'{{{NS_R}}}embed', rid)
    stretch = etree.SubElement(bf, f'{{{NS_A}}}stretch')
    etree.SubElement(stretch, f'{{{NS_A}}}fillRect')
    spPr = etree.SubElement(pic, f'{{{NS_XDR}}}spPr')
    xfrm = etree.SubElement(spPr, f'{{{NS_A}}}xfrm')
    off = etree.SubElement(xfrm, f'{{{NS_A}}}off'); off.set('x','0'); off.set('y','0')
    ext2 = etree.SubElement(xfrm, f'{{{NS_A}}}ext'); ext2.set('cx', str(cx)); ext2.set('cy', str(cy))
    pg = etree.SubElement(spPr, f'{{{NS_A}}}prstGeom'); pg.set('prst','rect')
    etree.SubElement(pg, f'{{{NS_A}}}avLst')
    etree.SubElement(anchor, f'{{{NS_XDR}}}clientData')
    return anchor

def fit_image_size(img_w, img_h, max_cx, max_cy):
    """画像をmax bounds内に収めるEMUサイズを計算（アスペクト比維持）"""
    aspect = img_w / img_h
    # 高さ基準で収める
    cy = max_cy
    cx = int(cy * aspect)
    # 幅がはみ出す場合は幅基準に切り替え
    if cx > max_cx:
        cx = max_cx
        cy = int(cx / aspect)
    return cx, cy

def build_drawing(tmpl_drawing_bytes, tmpl_rels_bytes, images_list, all_files, media_prefix):
    drawing = etree.fromstring(tmpl_drawing_bytes)
    rels    = etree.fromstring(tmpl_rels_bytes)
    for anchor in list(drawing):
        fe  = anchor.find(f'{{{NS_XDR}}}from')
        pic = anchor.find(f'{{{NS_XDR}}}pic')
        if fe is None or pic is None: continue
        if int(fe.find(f'{{{NS_XDR}}}row').text) == 45:
            ne = pic.find(f'.//{{{NS_XDR}}}cNvPr')
            if ne is not None and ne.get('name','') == '図 8':
                drawing.remove(anchor)
    # 全画像からオブジェクト（本体・台座等）を切り出し
    all_objects = []
    for img_data, in images_list:
        all_objects.extend(extract_objects(img_data))
    if not all_objects:
        return (etree.tostring(drawing, xml_declaration=True, encoding='UTF-8', standalone=True),
                etree.tostring(rels,    xml_declaration=True, encoding='UTF-8', standalone=True))
    n = len(all_objects)
    total_gap = IMG_GAP_EMU * (n - 1)
    per_obj_max_cx = (MAX_CX - total_gap) // n
    cur_col_off_emu = 0
    for idx, (obj_data, img_w, img_h) in enumerate(all_objects):
        rid = f'rId{20+idx}'; img_id = 20+idx
        fname = f'{media_prefix}_{idx}.png'
        cx, cy = fit_image_size(img_w, img_h, per_obj_max_cx, MAX_CY)
        col = IMG_COL_START + (cur_col_off_emu // EMU_PER_COL)
        col_off = cur_col_off_emu % EMU_PER_COL
        rel = etree.SubElement(rels, f'{{{NS_PKG}}}Relationship')
        rel.set('Id', rid)
        rel.set('Type','http://schemas.openxmlformats.org/officeDocument/2006/relationships/image')
        rel.set('Target', f'../media/{fname}')
        drawing.append(make_pic_anchor(rid, img_id, f'design_{idx}',
                                       col, IMG_ROW_START, cx, cy, col_off))
        all_files[f'xl/media/{fname}'] = obj_data
        cur_col_off_emu += cx + IMG_GAP_EMU
    return (etree.tostring(drawing, xml_declaration=True, encoding='UTF-8', standalone=True),
            etree.tostring(rels,    xml_declaration=True, encoding='UTF-8', standalone=True))

def fill_sheet(ws, d):
    tomorrow = datetime.now() + timedelta(days=1)
    ws['AD4'] = tomorrow.year; ws['AH4'] = tomorrow.month; ws['AK4'] = tomorrow.day
    ws['F11'] = d['nouki']
    ws['A22'].value = d['anken_stripped']
    all_items = [
        {'name':'本納品','qty':d['honnohin'],'unit':'個','price':d['tanka']},
        {'name':'予備',  'qty':d['yubi'],    'unit':'個','price':f"=Y{SLOTS[0]}"},
    ]
    for bi in d['betto_items']:
        all_items.append({'name':bi['name'],'qty':bi['qty'],'unit':bi['unit'],'price':bi['price']})
    for i, slot_row in enumerate(SLOTS):
        if i < len(all_items):
            item = all_items[i]
            has_name = item['name'] is not None and str(item['name']).strip() != ''
            ws.cell(row=slot_row, column=1).value  = item['name'] if has_name else None
            ws.cell(row=slot_row, column=16).value = item['qty'] if has_name else None
            ws.cell(row=slot_row, column=22).value = item['unit'] if has_name else None
            ws.cell(row=slot_row, column=25).value = (f"=Y{SLOTS[0]}" if i==1 else item['price']) if has_name else None
        else:
            ws.cell(row=slot_row, column=1).value  = None
            ws.cell(row=slot_row, column=16).value = None
            ws.cell(row=slot_row, column=22).value = None
            ws.cell(row=slot_row, column=25).value = None
    specs = [
        (46, f'■種別：{d["sku"]}種 '),
        (48, f'■{d["size_label"]}：{d["size_str"]}'),
        (50, '■材質：アクリル'),
        (52, f'■印刷：{d["insatsu"]}　 '),
        (54, '■個装：OPP'),
        (56, '■備考：以上の内容で進行いたします。'),
    ]
    for row, text in specs:
        cell = ws.cell(row=row, column=2)
        cell.value = text
        cell.font  = Font(name='游ゴシック', size=11)
        cell.alignment = Alignment(vertical='center')

def generate(xlsx_paths, template_path, output_path, orig_paths=None):
    tomorrow = datetime.now() + timedelta(days=1)
    shutil.copy(template_path, output_path)
    wb_out = load_workbook(output_path)
    with zipfile.ZipFile(template_path) as zt:
        tmpl_drawing = zt.read('xl/drawings/drawing1.xml')
        tmpl_rels    = zt.read('xl/drawings/_rels/drawing1.xml.rels')
    datasets = []
    for i, xlsx_path in enumerate(xlsx_paths):
        d = read_estimate(xlsx_path)
        img_path = orig_paths[i] if orig_paths else xlsx_path
        raw = get_images_from_sheet(img_path, '種別')
        datasets.append({'d': d, 'images': raw, 'path': xlsx_path})
    for idx, ds in enumerate(datasets):
        sheet_name = ds['d']['anken'][:31]
        if idx == 0:
            ws = wb_out['例']; ws.title = sheet_name
        else:
            ws = wb_out.copy_worksheet(wb_out.worksheets[0]); ws.title = sheet_name
        fill_sheet(ws, ds['d'])
    wb_out.save(output_path)
    with zipfile.ZipFile(output_path,'r') as zin:
        all_files = {name: zin.read(name) for name in zin.namelist()}
    with zipfile.ZipFile(template_path) as zt:
        for f in zt.namelist():
            if f.startswith('xl/media/'): all_files[f] = zt.read(f)
    for idx, ds in enumerate(datasets):
        drw_xml, drw_rels = build_drawing(tmpl_drawing, tmpl_rels,
                                          ds['images'], all_files, f'design_est{idx}')
        all_files[f'xl/drawings/drawing{idx+1}.xml'] = drw_xml
        all_files[f'xl/drawings/_rels/drawing{idx+1}.xml.rels'] = drw_rels
        s = all_files[f'xl/worksheets/sheet{idx+1}.xml'].decode('utf-8')
        if '<drawing' not in s:
            tag = '<drawing r:id="rId1" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/>'
            s = s.replace('</worksheet>', f'{tag}</worksheet>')
            all_files[f'xl/worksheets/sheet{idx+1}.xml'] = s.encode('utf-8')
        if idx > 0:
            all_files[f'xl/worksheets/_rels/sheet{idx+1}.xml.rels'] = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="/xl/drawings/drawing{idx+1}.xml" Id="rId1"/>
</Relationships>'''.encode('utf-8')
            ct = etree.fromstring(all_files['[Content_Types].xml'])
            if not any(f'drawing{idx+1}' in (el.get('PartName') or '') for el in ct):
                ov = etree.SubElement(ct, f'{{{NS_CT}}}Override')
                ov.set('PartName', f'/xl/drawings/drawing{idx+1}.xml')
                ov.set('ContentType','application/vnd.openxmlformats-officedocument.drawing+xml')
            all_files['[Content_Types].xml'] = etree.tostring(ct, xml_declaration=True, encoding='UTF-8', standalone=True)
            wb_xml = etree.fromstring(all_files['xl/workbook.xml'])
            dns = wb_xml.find(f'{{{NS_S}}}definedNames')
            if dns is None: dns = etree.SubElement(wb_xml, f'{{{NS_S}}}definedNames')
            if not any(dn.get('localSheetId') == str(idx) for dn in dns):
                dn = etree.SubElement(dns, f'{{{NS_S}}}definedName')
                dn.set('name','_xlnm.Print_Area'); dn.set('localSheetId', str(idx))
                dn.text = f"'{ds['d']['anken'][:31]}'!$A$1:$AM$60"
            all_files['xl/workbook.xml'] = etree.tostring(wb_xml, xml_declaration=True, encoding='UTF-8', standalone=True)
    with zipfile.ZipFile(output_path,'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in all_files.items():
            zout.writestr(name, data)
    return output_path

# ============================================================
# UI
# ============================================================
st.markdown("""
<div class="step-card">
  <span class="step-num">1</span><span class="step-title">見積算出表をアップロード（複数OK）</span>
</div>
""", unsafe_allow_html=True)

if 'xlsb_key' not in st.session_state:
    st.session_state.xlsb_key = 0

uploaded_xlsb = st.file_uploader(
    "見積算出表 .xlsb をドラッグ＆ドロップ",
    type=['xlsb'],
    accept_multiple_files=True,
    label_visibility="collapsed",
    key=f"xlsb_{st.session_state.xlsb_key}"
)

st.markdown("""
<div class="step-card">
  <span class="step-num">2</span><span class="step-title">テンプレートをアップロード</span>
</div>
""", unsafe_allow_html=True)

uploaded_template = st.file_uploader(
    "最終見積書テンプレート .xlsx",
    type=['xlsx'],
    label_visibility="collapsed",
    key="template"
)

st.markdown("<br>", unsafe_allow_html=True)

if uploaded_xlsb and uploaded_template:
    names = [f.name for f in uploaded_xlsb]
    st.markdown(f"""
    <div class="warning-box">
    📋 <b>{len(uploaded_xlsb)}件</b> の見積算出表を受け取りました：{"　".join(names)}
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    if st.button("📄 最終見積書を生成する", type="primary", use_container_width=True):
        with st.spinner("生成中... しばらくお待ちください"):
            with tempfile.TemporaryDirectory() as tmpdir:
                try:
                    # xlsb → xlsx 変換（pyxlsb使用）
                    xlsx_paths = []
                    xlsb_orig_paths = []
                    progress = st.progress(0)
                    for i, uf in enumerate(uploaded_xlsb):
                        import pyxlsb, io as _io
                        from openpyxl import Workbook as _WB
                        xlsb_bytes = uf.read()
                        # 元xlsbを保存（画像抽出用）
                        xlsb_orig_path = os.path.join(tmpdir, uf.name)
                        with open(xlsb_orig_path, 'wb') as f_out:
                            f_out.write(xlsb_bytes)
                        xlsb_orig_paths.append(xlsb_orig_path)
                        xlsx_path = os.path.join(tmpdir, uf.name.replace('.xlsb', '.xlsx'))
                        with pyxlsb.open_workbook(_io.BytesIO(xlsb_bytes)) as wb_xlsb:
                            wb_new = _WB()
                            wb_new.remove(wb_new.active)
                            for si, sname in enumerate(wb_xlsb.sheets):
                                title = sname.strip() if sname and sname.strip() else f'Sheet{si+1}'
                                ws_new = wb_new.create_sheet(title=title)
                                with wb_xlsb.get_sheet(sname) as sheet:
                                    for row in sheet.rows():
                                        for cell in row:
                                            if cell.v is not None:
                                                ws_new.cell(row=cell.r + 1, column=cell.c + 1, value=cell.v)
                            wb_new.save(xlsx_path)
                        xlsx_paths.append(xlsx_path)
                        progress.progress((i+1)/len(uploaded_xlsb))

                    # テンプレート保存
                    template_path = os.path.join(tmpdir, '_template.xlsx')
                    with open(template_path, 'wb') as out:
                        out.write(uploaded_template.read())

                    # 生成
                    today = datetime.now().strftime('%Y%m%d')
                    output_path = os.path.join(tmpdir, f'最終見積書_カードラボ様_{today}.xlsx')
                    generate(xlsx_paths, template_path, output_path, orig_paths=xlsb_orig_paths)

                    with open(output_path, 'rb') as f:
                        output_bytes = f.read()

                    fname = f'最終見積書_カードラボ様_{today}.xlsx'

                    # 結果をsession_stateに保存してリセット
                    st.session_state.result_bytes = output_bytes
                    st.session_state.result_fname = fname
                    st.session_state.xlsb_key += 1
                    st.rerun()

                except Exception as e:
                    st.error(f"エラーが発生しました：{str(e)}")

else:
    st.markdown("""
    <div style="text-align:center; color:#c8a0b8; padding: 32px 0; font-size:0.9rem;">
        🌸 見積算出表とテンプレートをアップロードしてください
    </div>
    """, unsafe_allow_html=True)

# 生成結果の表示（rerun後）
if 'result_bytes' in st.session_state and st.session_state.result_bytes is not None:
    import base64 as _b64
    _mascot_img = Image.open('mascot.png')
    _mascot_img.thumbnail((150, 220), Image.LANCZOS)
    _buf = io.BytesIO()
    _mascot_img.save(_buf, format='PNG')
    _mascot_b64 = _b64.b64encode(_buf.getvalue()).decode()

    st.markdown(f"""
    <div class="success-box" style="display:flex;align-items:center;justify-content:center;gap:20px;">
        <img src="data:image/png;base64,{_mascot_b64}" style="height:200px;margin:-30px 0;">
        <span style="font-size:1.2rem;">🎉 <b>生成できました！</b></span>
    </div>
    """, unsafe_allow_html=True)

    st.download_button(
        label=f"⬇️  {st.session_state.result_fname}  を\nダウンロード",
        data=st.session_state.result_bytes,
        file_name=st.session_state.result_fname,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        use_container_width=True
    )
    # ダウンロード後にクリア
    st.session_state.result_bytes = None
    st.session_state.result_fname = None

# フッター
st.markdown("<hr style='border:none;border-top:1px solid #f0d4e4;margin-top:32px;'>", unsafe_allow_html=True)
st.markdown("<p style='text-align:center;color:#c8a0b8;font-size:0.8rem;'>株式会社アイナック 事業推進部（東京）</p>",
            unsafe_allow_html=True)
